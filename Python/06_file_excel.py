# ============================================================================
# 6. FILE HANDLING - EXCEL
# ============================================================================
# WHAT: Create and manipulate Excel files using openpyxl library
# WHY: Excel files are widely used in business, better formatting than CSV
# WHEN: Need formatted reports, multiple sheets, formulas

# NOTE: Requires openpyxl library - install with: pip install openpyxl

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill

    print("="*60)
    print("EXCEL OPERATIONS WITH OPENPYXL")
    print("="*60)

    # ============================================================================
    # CREATING EXCEL FILES
    # ============================================================================

    print("\n1. Creating Excel File:")

    # Create new workbook and get active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "People"  # Name the sheet

    # Add rows to sheet
    ws.append(["Name", "Age", "Department"])  # Header row
    ws.append(["Alice", 30, "Engineering"])
    ws.append(["Bob", 25, "Sales"])
    ws.append(["Carol", 35, "Marketing"])

    # Save workbook to file
    wb.save("people.xlsx")
    print("✓ Excel file created: people.xlsx")

    # ============================================================================
    # READING EXCEL FILES
    # ============================================================================

    print("\n2. Reading Excel File:")

    # Load existing workbook
    wb = load_workbook("people.xlsx")
    ws = wb.active

    print(f"Sheet name: {ws.title}")
    print("Contents:")

    for row in ws.iter_rows(values_only=True):
        print(f"  {row}")

    # ============================================================================
    # ACCESSING SPECIFIC CELLS
    # ============================================================================

    print("\n3. Accessing Specific Cells:")

    # Access cell by coordinate
    cell_a1 = ws["A1"]
    print(f"Cell A1: {cell_a1.value}")

    # Access cell by row/column
    cell = ws.cell(row=2, column=1)
    print(f"Cell at row 2, col 1: {cell.value}")

    # ============================================================================
    # EDITING CELLS
    # ============================================================================

    print("\n4. Editing Cells:")

    # Modify cell value
    ws["A2"] = "Alicia"  # Change Alice to Alicia
    ws.cell(row=3, column=2, value=26)  # Change Bob's age

    # Add new row
    ws.append(["David", 28, "HR"])

    print("✓ Modified cells and added new row")

    # ============================================================================
    # FORMATTING CELLS
    # ============================================================================

    print("\n5. Formatting Cells (Colors, Fonts):")

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Color a cell
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    print("✓ Applied formatting")

    # ============================================================================
    # WORKING WITH FORMULAS
    # ============================================================================

    print("\n6. Working with Formulas:")

    # Create new sheet for formulas
    ws_calc = wb.create_sheet("Calculations")
    ws_calc.append(["Item", "Price", "Quantity", "Total"])
    ws_calc.append(["Apple", 1.5, 10, "=B2*C2"])
    ws_calc.append(["Banana", 0.8, 15, "=B3*C3"])
    ws_calc.append(["Total", "", "", "=SUM(D2:D3)"])

    print("✓ Added formulas to 'Calculations' sheet")

    # ============================================================================
    # SAVING CHANGES
    # ============================================================================

    # Save all changes
    wb.save("people.xlsx")
    print("\n✓ All changes saved to people.xlsx")

    print("""
    💡 Excel Automation Use Cases (ATBS):
    - Generate reports from data
    - Bulk update spreadsheets
    - Extract data from multiple Excel files
    - Create invoices, receipts, forms
    - Merge data from CSV → Excel
    - Create dashboards with formulas
    """)

    print("\n📊 For BI Integration:")
    print("- Read Excel → Transform → Load to Power BI")
    print("- Export processed data to Excel for stakeholders")
    print("- Automate monthly/weekly report generation")

    # Clean up
    import os
    if os.path.exists("people.xlsx"):
        os.remove("people.xlsx")
        print("\nCleaned up: people.xlsx")

except ImportError:
    print("openpyxl not installed")
    print("Install with: pip install openpyxl")
